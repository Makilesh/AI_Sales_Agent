
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from models.lead import Lead


def export_to_excel(
    leads: list[Lead], 
    qualifications: list[dict], 
    filename: str
) -> None:
    """
    Export qualified leads to Excel with formatting.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        filename: Output Excel file path
    """
    # Handle empty leads gracefully
    if not leads or not qualifications:
        print(f"⚠️  No qualified leads to export")
        # Create empty workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualified Leads"

        headers = [
            "Author", "Source", "Subreddit", "Post Type", "Content", "URL",
            "Engagement Score", "Targeted Search", "Search Phrase",
            "Is Qualified", "Confidence", "LLM Provider", "Reason", "Service Match", "Timestamp"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filename)
        return
    
    if len(leads) != len(qualifications):
        raise ValueError(f"Leads ({len(leads)}) and qualifications ({len(qualifications)}) must have same length")
    
    # Ensure directory exists
    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Qualified Leads"
    
    # Define headers (IMPROVED: Added metadata columns for Reddit)
    headers = [
        "Author",
        "Source",
        "Subreddit",           # NEW: Reddit subreddit name
        "Post Type",           # NEW: post/comment/submission
        "Content",
        "URL",
        "Engagement Score",
        "Targeted Search",     # NEW: High-intent search flag
        "Search Phrase",       # NEW: Which search phrase matched
        "Is Qualified",
        "Confidence",
        "LLM Provider",        # NEW: openai/gemini fallback tracking
        "Reason",
        "Service Match",
        "Timestamp"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Combine leads with qualifications and sort by confidence score descending
    combined = list(zip(leads, qualifications))
    combined.sort(key=lambda x: x[1].get('confidence_score', 0.0), reverse=True)
    
    # Define fills for conditional formatting
    qualified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    not_qualified_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    search_lead_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Gold/yellow for targeted search leads (HIGHEST PRIORITY)
    
    # Write data rows
    for row_idx, (lead, qual) in enumerate(combined, 2):
        # Truncate content to 200 chars
        content = lead.content[:200] + "..." if len(lead.content) > 200 else lead.content
        
        # Format service match as comma-separated string
        service_match = ", ".join(qual.get('service_match', []))
        
        # Format timestamp
        timestamp_str = lead.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        
        # Prepare row data (IMPROVED: Added metadata columns)
        row_data = [
            lead.author,
            lead.source,
            lead.subreddit or "N/A",                                    # NEW: Subreddit
            lead.metadata.get('post_type', 'N/A'),                      # NEW: Post type
            content,
            lead.url,
            lead.engagement_score,
            "Yes" if lead.metadata.get('targeted_search') else "No",   # NEW: Search flag
            lead.metadata.get('search_phrase', 'N/A'),                  # NEW: Search phrase
            "Yes" if qual.get('is_qualified', False) else "No",
            round(qual.get('confidence_score', 0.0), 2),
            qual.get('llm_provider', 'N/A'),                            # NEW: LLM provider
            qual.get('reason', ''),
            service_match,
            timestamp_str
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in [5, 13]))  # Wrap text for Content (col 5) and Reason (col 13)

            # IMPROVED (ISSUE #3): Apply conditional formatting with search lead priority
            # Priority: Targeted Search (gold) > Qualified (green) > Not Qualified (red)
            if lead.metadata.get('targeted_search'):
                # Gold highlight for targeted search leads (HIGHEST PRIORITY)
                cell.fill = search_lead_fill
            elif qual.get('is_qualified', False):
                cell.fill = qualified_fill
            else:
                cell.fill = not_qualified_fill
    
    # Auto-adjust column widths (IMPROVED: Updated for new columns)
    column_widths = {
        1: 20,   # Author
        2: 12,   # Source
        3: 18,   # Subreddit (NEW)
        4: 12,   # Post Type (NEW)
        5: 50,   # Content
        6: 40,   # URL
        7: 15,   # Engagement Score
        8: 15,   # Targeted Search (NEW)
        9: 30,   # Search Phrase (NEW)
        10: 12,  # Is Qualified
        11: 12,  # Confidence
        12: 12,  # LLM Provider (NEW)
        13: 50,  # Reason
        14: 30,  # Service Match
        15: 20   # Timestamp
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(filename)
    
    qualified_count = sum(1 for _, q in combined if q.get('is_qualified', False))
    search_lead_count = sum(1 for l, _ in combined if l.metadata.get('targeted_search'))
    print(f"\n✅ Exported {len(leads)} leads to {filename}")
    print(f"   • Qualified: {qualified_count}")
    print(f"   • Not Qualified: {len(leads) - qualified_count}")
    if search_lead_count > 0:
        print(f"   • 🎯 Targeted Search Leads: {search_lead_count} (highlighted in gold)")
    print(f"   • Sorted by confidence score (highest first)")


def export_qualified_only(
    leads: list[Lead],
    qualifications: list[dict],
    filename: str,
    min_confidence: float = 0.0
) -> None:
    """
    Export only qualified leads to Excel.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        filename: Output Excel file path
        min_confidence: Minimum confidence score to include (default: 0.0)
    """
    # Filter to qualified leads only
    filtered = [
        (lead, qual) 
        for lead, qual in zip(leads, qualifications)
        if qual.get('is_qualified', False) and qual.get('confidence_score', 0.0) >= min_confidence
    ]
    
    if not filtered:
        print(f"⚠️  No qualified leads found with confidence >= {min_confidence}")
        return
    
    qualified_leads, qualified_quals = zip(*filtered)
    
    print(f"📊 Exporting {len(qualified_leads)} qualified leads (confidence >= {min_confidence})...")
    export_to_excel(list(qualified_leads), list(qualified_quals), filename)


def export_by_service(
    leads: list[Lead],
    qualifications: list[dict],
    service: str,
    filename: str,
    min_confidence: float = 0.0
) -> None:
    """
    Export leads matching specific service to Excel.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        service: Service to filter by ('RWA', 'Crypto/Blockchain', 'AI/ML')
        filename: Output Excel file path
        min_confidence: Minimum confidence score to include (default: 0.0)
    """
    # Filter to specific service
    filtered = [
        (lead, qual)
        for lead, qual in zip(leads, qualifications)
        if qual.get('is_qualified', False)
        and service in qual.get('service_match', [])
        and qual.get('confidence_score', 0.0) >= min_confidence
    ]
    
    if not filtered:
        print(f"⚠️  No qualified leads found for service '{service}' with confidence >= {min_confidence}")
        return
    
    service_leads, service_quals = zip(*filtered)
    
    print(f"📊 Exporting {len(service_leads)} {service} leads (confidence >= {min_confidence})...")
    export_to_excel(list(service_leads), list(service_quals), filename)


def export_all_leads_to_excel(leads: list[Lead], filename: str) -> None:
    """
    Export ALL leads (qualified + unqualified) to Excel.
    Useful when you want raw scraped data in Excel format without AI qualification.
    
    Args:
        leads: List of Lead objects
        filename: Output Excel file path
    """
    # Handle empty leads gracefully
    if not leads:
        print(f"⚠️  No leads to export")
        # Create empty workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "All Leads"
        
        headers = ["Author", "Source", "Content", "URL", "Engagement Score", "Timestamp"]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filename)
        print(f"✅ Created empty Excel file: {filename}")
        return
    
    # Ensure directory exists
    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"
    
    # Define headers (simpler than qualified version)
    headers = ["Author", "Source", "Content", "URL", "Engagement Score", "Timestamp"]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sort by engagement score (most engagement first)
    sorted_leads = sorted(leads, key=lambda x: x.engagement_score or 0, reverse=True)
    
    # Write data rows
    for row_idx, lead in enumerate(sorted_leads, 2):
        # Truncate content to 300 chars (more space since fewer columns)
        content = lead.content[:300] + "..." if len(lead.content) > 300 else lead.content
        
        # Format timestamp
        timestamp_str = lead.timestamp.strftime("%Y-%m-%d %H:%M:%S") if lead.timestamp else ""
        
        # Prepare row data
        row_data = [
            lead.author,
            lead.source,
            content,
            lead.url,
            lead.engagement_score or 0,
            timestamp_str
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 3))  # Wrap text for Content
    
    # Auto-adjust column widths
    column_widths = {
        1: 20,   # Author
        2: 12,   # Source
        3: 60,   # Content
        4: 50,   # URL
        5: 18,   # Engagement Score
        6: 20    # Timestamp
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(filename)
    
    print(f"✅ Excel export complete: {filename}")
    print(f"   • Total leads: {len(leads)}")
    print(f"   • Sorted by engagement score (highest first)")
